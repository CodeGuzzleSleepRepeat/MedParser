from bs4 import BeautifulSoup
import urllib3
import pandas as pd
import xlrd

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, flag=False, 
                       **to_excel_kwargs):
   
    from openpyxl import load_workbook

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        writer.book = load_workbook(filename)

        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs, header = flag)

    writer.save()



def parse(soup, link, names, cur_flag):
    c = 0
    k2 = 0
    mem = ""
    len = 17
    flag = True
    data = pd.DataFrame(columns = names, index = range(8))
    for line in soup.recursiveChildGenerator():
        if line.name == "th":
            k2 = 1
        if k2 == 1 and line.name == "td":
            if line.text == check2 and mem == check1:
                flag = True
                break
            data[names[c % len]][int (c / len)] = line.text
            c = c + 1
            mem = line.text
    if c % len != 0:
        flag = False
    del data['N']
    data = data[ : int(c / len)]
    append_df_to_excel(link, data, sheet_name = 'Sheet1', index = False, flag = cur_flag)
    return flag

link = "MedResult2.xlsx"
workbook = xlrd.open_workbook(link)
worksheet = workbook.sheet_by_index(0)
check1 = worksheet.cell(2, 0).value
check2 = worksheet.cell(2, 1).value

gotData = pd.read_excel(link, sheet_name = 'Sheet1')

agent = {'user-agent' : 'Chrome'}
http = urllib3.PoolManager(headers = agent)

doc = http.request('GET', 'http://grls.rosminzdrav.ru/CiPermitionReg.aspx?PermYear=0&DateInc=&NumInc=&DateBeg=&DateEnd=&Protocol=&RegNm=&Statement=&ProtoNum=&idCIStatementCh=&Qualifier=&CiPhase=&RangeOfApp=&Torg=&LFDos=&Producer=&Recearcher=&sponsorCountry=&MedBaseCount=&CiType=&PatientCount=&OrgDocOut=2&Status=1%2c2%2c3%2c4&NotInReg=0&All=0&PageSize=8&order=date_perm&orderType=desc&pagenum=1')

soup = BeautifulSoup(doc.data, features = "lxml")

k = 0

names = []

names.append("N")
if soup.th == None:
    print("Data is temporarily unavailable")
    exit(0)
for line in soup.recursiveChildGenerator():
    if line.name == "th":
        k = 1
    if k == 1 and line.name == "a":
        names.append(line.text)
        k = 0

buf = pd.DataFrame()
buf.to_excel(link)
flag = parse(soup, link, names, True)

i = 2

while flag:
    doc = http.request('GET', 'http://grls.rosminzdrav.ru/CiPermitionReg.aspx?PermYear=0&DateInc=&Num' +
        'Inc=&DateBeg=&DateEnd=&Protocol=&RegNm=&Statement=&ProtoNum=&idCIStatementCh=&Qualifier=&CiPhase=&RangeOfApp=&Torg=&' +
        'LFDos=&Producer=&Recearcher=&sponsorCountry=&MedBaseCount=&CiType=&PatientCount=&OrgDocOut=2&Status=1%2c2%2c3%2c4&NotInReg=0&' +
        'All=0&PageSize=8&order=date_perm&orderType=desc&pagenum=' + str(i))
    soup = BeautifulSoup(doc.data, features = "lxml")
    if soup.th == None:
        break
    flag = parse(soup, link, names, False)
    i = i + 1
append_df_to_excel(link, gotData[1:], sheet_name = 'Sheet1', index = False, flag = False)
